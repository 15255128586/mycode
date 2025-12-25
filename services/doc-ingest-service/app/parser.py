import io
from typing import Any, Dict, List

import pdfplumber
from docx import Document
from openpyxl import load_workbook


def _append_segment(segments: List[Dict[str, Any]], text: str, source: Dict[str, Any]) -> None:
    cleaned = text.strip()
    if not cleaned:
        return
    segments.append({"text": cleaned, "source": source})


def parse_markdown(data: bytes) -> List[Dict[str, Any]]:
    text = data.decode("utf-8", errors="replace")
    segments: List[Dict[str, Any]] = []
    for idx, line in enumerate(text.splitlines(), start=1):
        _append_segment(segments, line, {"line": idx})
    return segments


def parse_docx(data: bytes) -> List[Dict[str, Any]]:
    document = Document(io.BytesIO(data))
    segments: List[Dict[str, Any]] = []
    for idx, paragraph in enumerate(document.paragraphs, start=1):
        _append_segment(segments, paragraph.text, {"paragraph": idx})
    return segments


def parse_xlsx(data: bytes) -> List[Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    segments: List[Dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(cell) for cell in row if cell is not None]
            if not values:
                continue
            text = "\t".join(values)
            _append_segment(segments, text, {"sheet": sheet.title, "row": row_idx})
    return segments


def parse_pdf(data: bytes) -> List[Dict[str, Any]]:
    segments: List[Dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line_idx, line in enumerate(text.splitlines(), start=1):
                _append_segment(segments, line, {"page": page_idx, "line": line_idx})
    return segments


def parse_file(filename: str, data: bytes) -> List[Dict[str, Any]]:
    lowered = filename.lower()
    if lowered.endswith(".md") or lowered.endswith(".markdown"):
        return parse_markdown(data)
    if lowered.endswith(".docx"):
        return parse_docx(data)
    if lowered.endswith(".xlsx"):
        return parse_xlsx(data)
    if lowered.endswith(".pdf"):
        return parse_pdf(data)
    raise ValueError("unsupported_file_type")
